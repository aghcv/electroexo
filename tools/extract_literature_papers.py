#!/usr/bin/env python3
"""Extract table-linked paper references from the literature evidence workbook."""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


TABLE_RE = re.compile(r"\bA\s*(\d{1,2})\b", re.IGNORECASE)


def clean(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def table_id(*values: object) -> str | None:
    for value in values:
        text = clean(value)
        match = TABLE_RE.search(text)
        if match:
            return f"A{int(match.group(1)):02d}"
    return None


def main() -> int:
    if len(sys.argv) != 2:
        print("usage: extract_literature_papers.py workbook.xlsx", file=sys.stderr)
        return 2

    workbook_path = Path(sys.argv[1]).expanduser()
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    records: list[dict[str, object]] = []

    for ws in wb.worksheets:
        active_table: str | None = None
        for row in ws.iter_rows():
            values = [clean(cell.value) for cell in row]
            row_table = table_id(*values[:11])
            if row_table:
                active_table = row_table

            for idx in range(11, min(16, len(values))):
                text = values[idx]
                if not text:
                    continue
                cell_table = table_id(text) or active_table
                if not cell_table:
                    continue
                records.append(
                    {
                        "sheet": ws.title,
                        "row": row[0].row,
                        "column": chr(ord("A") + idx),
                        "table": cell_table,
                        "text": text,
                    }
                )

    unique: dict[tuple[str, str], dict[str, object]] = {}
    for record in records:
        key = (str(record["table"]), str(record["text"]).lower())
        unique.setdefault(key, record)

    print(json.dumps(list(unique.values()), indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
