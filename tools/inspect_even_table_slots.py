#!/usr/bin/env python3
"""Print even-table paper slots from the literature tracker."""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook


def main() -> int:
    workbook = Path(sys.argv[1])
    wb = load_workbook(workbook, data_only=False, read_only=True)
    ws = wb["Literature Tracker"]
    print("max", ws.max_row, ws.max_column)
    print([ws.cell(1, c).value for c in range(1, 17)])
    for row in range(2, ws.max_row + 1):
        table = str(ws.cell(row, 1).value or "")
        if not (table.startswith("A") and table[1:].isdigit() and int(table[1:]) % 2 == 0):
            continue
        empties = [chr(64 + col) for col in range(12, 17) if ws.cell(row, col).value in (None, "")]
        papers = [ws.cell(row, col).value for col in range(12, 17)]
        print(row, table, ws.cell(row, 2).value, "empties", empties, "papers", papers)
        if empties:
            print("  layer:", ws.cell(row, 3).value)
            print("  module:", ws.cell(row, 4).value)
            print("  submodule:", ws.cell(row, 5).value)
            print("  form:", ws.cell(row, 6).value)
            print("  focus:", ws.cell(row, 7).value)
            print("  relation:", ws.cell(row, 9).value)
            print("  inputs:", ws.cell(row, 10).value)
            print("  outputs:", ws.cell(row, 11).value)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
