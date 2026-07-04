#!/usr/bin/env python3
"""Collect workbook-cited PDFs into table-number folders."""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import shutil
import sqlite3
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook


TABLE_RE = re.compile(r"\bA\s*(\d{1,2})\b", re.IGNORECASE)
DOI_RE = re.compile(r"\b10\.\d{4,9}/[^\s;,\]\)\"<>]+", re.IGNORECASE)
PMCID_RE = re.compile(r"\bPMC\d+\b", re.IGNORECASE)
PMID_RE = re.compile(r"\bPMID:\s*(\d+)", re.IGNORECASE)
YEAR_RE = re.compile(r"\b(19|20)\d{2}\b")


def clean(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def safe_name(value: str, limit: int = 140) -> str:
    value = re.sub(r"[/\\:*?\"<>|]", "_", value)
    value = re.sub(r"\s+", " ", value).strip(" .")
    return value[:limit].rstrip(" .") or "paper"


def norm(value: str) -> str:
    value = value.lower()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def table_id(*values: object) -> str | None:
    for value in values:
        match = TABLE_RE.search(clean(value))
        if match:
            return f"A{int(match.group(1)):02d}"
    return None


def doi_from_text(text: str) -> str:
    match = DOI_RE.search(text)
    if not match:
        return ""
    return match.group(0).rstrip(".").lower()


def pmcid_from_text(text: str) -> str:
    match = PMCID_RE.search(text)
    return match.group(0).upper() if match else ""


def pmid_from_text(text: str) -> str:
    match = PMID_RE.search(text)
    return match.group(1) if match else ""


def title_from_ref(text: str) -> str:
    quoted = re.findall(r"[“\"]([^”\"]{20,240})[”\"]", text)
    if quoted:
        return quoted[0]

    apa = re.search(r"\((?:19|20)\d{2}[a-z]?\)\.\s+(.+?)\.\s+[A-Z][A-Za-z& ]+,", text)
    if apa:
        return apa.group(1).strip()

    pubmed = re.search(r"\.\s+([^.;]{20,260}?)\.\s+[A-Z][A-Za-z .&()'-]+\.?\s+(?:19|20)\d{2}", text)
    if pubmed:
        return pubmed.group(1).strip()

    parts = re.split(r"\.\s+", text)
    for part in parts[1:5]:
        candidate = part.strip()
        if len(candidate) > 18 and not re.match(r"^(doi|pmid|pmcid)\b", candidate, re.I):
            if not re.search(r"\b(et al|vol|no|pp|e?pub)\b", candidate, re.I):
                return candidate.rstrip(".")
    return parts[0][:120].strip()


def year_from_text(text: str) -> str:
    matches = YEAR_RE.findall(text)
    return matches[0] if matches else "undated"


@dataclass
class Paper:
    table: str
    text: str
    sheet: str
    row: int
    column: str
    title: str = ""
    doi: str = ""
    pmcid: str = ""
    pmid: str = ""
    year: str = ""
    sources: list[str] = field(default_factory=list)
    pdf: str = ""
    status: str = "pending"

    @property
    def key(self) -> str:
        if self.doi:
            return f"doi:{self.doi}"
        if self.pmcid:
            return f"pmcid:{self.pmcid}"
        if self.pmid:
            return f"pmid:{self.pmid}"
        return f"title:{norm(self.title or self.text)[:120]}"

    @property
    def basename(self) -> str:
        bits = [self.year, self.title or self.text[:80]]
        if self.doi:
            bits.append(self.doi.replace("/", "_"))
        elif self.pmcid:
            bits.append(self.pmcid)
        digest = hashlib.sha1(self.key.encode()).hexdigest()[:8]
        return safe_name(" - ".join(bits), 165) + f" - {digest}.pdf"


def extract_papers(workbook_path: Path) -> list[Paper]:
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    papers: dict[tuple[str, str], Paper] = {}
    for ws in wb.worksheets:
        active_table: str | None = None
        for row in ws.iter_rows():
            values = [clean(cell.value) for cell in row]
            row_table = table_id(*values[:11])
            if row_table:
                active_table = row_table
            for idx in range(11, min(16, len(values))):
                text = values[idx]
                if not text:
                    continue
                if len(text) < 12 and not (DOI_RE.search(text) or PMCID_RE.search(text) or PMID_RE.search(text)):
                    continue
                current_table = table_id(text) or active_table
                if not current_table:
                    continue
                paper = Paper(
                    table=current_table,
                    text=text,
                    sheet=ws.title,
                    row=row[0].row,
                    column=chr(ord("A") + idx),
                )
                paper.title = title_from_ref(text)
                paper.doi = doi_from_text(text)
                paper.pmcid = pmcid_from_text(text)
                paper.pmid = pmid_from_text(text)
                paper.year = year_from_text(text)
                key = (paper.table, paper.key)
                if key not in papers:
                    papers[key] = paper
    return sorted(papers.values(), key=lambda p: (p.table, p.row, p.column, p.title.lower()))


def zotero_index(zotero_dir: Path) -> list[dict[str, str]]:
    db = zotero_dir / "zotero.sqlite"
    if not db.exists():
        return []
    db_copy = Path("/tmp") / f"zotero-{os.getpid()}.sqlite"
    shutil.copy2(db, db_copy)
    conn = sqlite3.connect(f"file:{db_copy}?mode=ro", uri=True)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        """
        SELECT ia.parentItemID AS parent_id, attach.key AS attachment_key, ia.path AS attachment_path,
               idv_title.value AS title, idv_doi.value AS doi
        FROM itemAttachments ia
        JOIN items attach ON attach.itemID = ia.itemID
        LEFT JOIN itemData id_title ON id_title.itemID = ia.parentItemID
        LEFT JOIN fields f_title ON f_title.fieldID = id_title.fieldID AND f_title.fieldName = 'title'
        LEFT JOIN itemDataValues idv_title ON idv_title.valueID = id_title.valueID
        LEFT JOIN itemData id_doi ON id_doi.itemID = ia.parentItemID
        LEFT JOIN fields f_doi ON f_doi.fieldID = id_doi.fieldID AND f_doi.fieldName = 'DOI'
        LEFT JOIN itemDataValues idv_doi ON idv_doi.valueID = id_doi.valueID
        WHERE ia.contentType = 'application/pdf'
        """
    ).fetchall()
    conn.close()
    try:
        db_copy.unlink()
    except OSError:
        pass

    out: list[dict[str, str]] = []
    for row in rows:
        path = row["attachment_path"] or ""
        if path.startswith("storage:"):
            pdf = zotero_dir / "storage" / row["attachment_key"] / path.removeprefix("storage:")
        else:
            pdf = Path(path).expanduser()
        if pdf and pdf.exists():
            out.append(
                {
                    "path": str(pdf),
                    "title": clean(row["title"]),
                    "title_norm": norm(clean(row["title"])),
                    "doi": clean(row["doi"]).lower(),
                }
            )
    return out


def copy_pdf(src: Path, dst: Path) -> bool:
    if dst.exists() and dst.stat().st_size > 1000:
        return True
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst)
    return dst.exists() and dst.stat().st_size > 1000


def find_zotero(paper: Paper, index: list[dict[str, str]]) -> Path | None:
    title_norm = norm(paper.title)
    for item in index:
        if paper.doi and item["doi"] and paper.doi in item["doi"]:
            return Path(item["path"])
    if len(title_norm) > 20:
        words = set(title_norm.split())
        for item in index:
            other = item["title_norm"]
            if title_norm and (title_norm in other or other in title_norm):
                return Path(item["path"])
            if len(words) >= 5:
                overlap = len(words & set(other.split()))
                if overlap / max(len(words), 1) >= 0.7:
                    return Path(item["path"])
    return None


def url_json(url: str, timeout: int = 30) -> object | None:
    req = urllib.request.Request(url, headers={"User-Agent": "codex-literature-pdf-collector/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=timeout) as response:
            return json.loads(response.read().decode("utf-8"))
    except Exception:
        return None


def download(url: str, dst: Path, timeout: int = 60) -> bool:
    if dst.exists() and dst.stat().st_size > 1000:
        return True
    req = urllib.request.Request(url, headers={"User-Agent": "codex-literature-pdf-collector/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=timeout) as response:
            data = response.read()
        if not data.startswith(b"%PDF") and b"/PDF" not in data[:4096]:
            return False
        dst.parent.mkdir(parents=True, exist_ok=True)
        dst.write_bytes(data)
        return dst.stat().st_size > 1000
    except (urllib.error.URLError, TimeoutError, OSError):
        return False


def online_urls(paper: Paper, email: str) -> list[str]:
    urls: list[str] = []
    if paper.pmcid:
        urls.append(f"https://www.ncbi.nlm.nih.gov/pmc/articles/{paper.pmcid}/pdf/")
    if paper.pmid:
        data = url_json(
            "https://www.ncbi.nlm.nih.gov/pmc/utils/idconv/v1.0/?"
            + urllib.parse.urlencode({"ids": paper.pmid, "format": "json"})
        )
        try:
            pmcid = data["records"][0].get("pmcid", "")
            if pmcid:
                urls.append(f"https://www.ncbi.nlm.nih.gov/pmc/articles/{pmcid}/pdf/")
        except Exception:
            pass
    if paper.doi:
        data = url_json(f"https://api.openalex.org/works/doi:{urllib.parse.quote(paper.doi)}?mailto={email}")
        try:
            for loc in data.get("locations", []) if isinstance(data, dict) else []:
                source = loc.get("source") or {}
                landing = loc.get("landing_page_url") or ""
                pdf = loc.get("pdf_url") or ""
                if pdf:
                    urls.append(pdf)
                if "ncbi.nlm.nih.gov/pmc" in landing:
                    match = PMCID_RE.search(landing)
                    if match:
                        urls.append(f"https://www.ncbi.nlm.nih.gov/pmc/articles/{match.group(0).upper()}/pdf/")
                if source.get("host_organization_name", "").lower() == "public library of science":
                    urls.append(f"https://journals.plos.org/plosone/article/file?id={paper.doi}&type=printable")
        except Exception:
            pass
        urls.append(f"https://doi.org/{paper.doi}")

    seen: set[str] = set()
    unique = []
    for url in urls:
        if url and url not in seen:
            unique.append(url)
            seen.add(url)
    return unique


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--papers-dir", required=True, type=Path)
    parser.add_argument("--zotero-dir", required=True, type=Path)
    parser.add_argument("--email", default="aghorban@odu.edu")
    parser.add_argument("--online", action="store_true")
    args = parser.parse_args()

    papers = extract_papers(args.workbook)
    zindex = zotero_index(args.zotero_dir)
    args.papers_dir.mkdir(parents=True, exist_ok=True)

    for paper in papers:
        table_dir = args.papers_dir / paper.table
        dst = table_dir / paper.basename
        zotero_pdf = find_zotero(paper, zindex)
        if zotero_pdf and copy_pdf(zotero_pdf, dst):
            paper.status = "zotero"
            paper.pdf = str(dst)
            paper.sources.append(str(zotero_pdf))
            continue
        if args.online:
            for url in online_urls(paper, args.email):
                if download(url, dst):
                    paper.status = "online"
                    paper.pdf = str(dst)
                    paper.sources.append(url)
                    time.sleep(0.25)
                    break
        if not paper.pdf:
            paper.status = "missing"

    manifest = [
        {
            "table": p.table,
            "status": p.status,
            "pdf": p.pdf,
            "title": p.title,
            "doi": p.doi,
            "pmcid": p.pmcid,
            "pmid": p.pmid,
            "sheet": p.sheet,
            "cell": f"{p.column}{p.row}",
            "sources": p.sources,
            "reference": p.text,
        }
        for p in papers
    ]
    (args.papers_dir / "manifest.json").write_text(json.dumps(manifest, indent=2, ensure_ascii=False))

    summary: dict[str, dict[str, int]] = {}
    for item in manifest:
        summary.setdefault(item["table"], {}).setdefault(item["status"], 0)
        summary[item["table"]][item["status"]] += 1
    print(json.dumps({"total": len(manifest), "summary": summary}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
