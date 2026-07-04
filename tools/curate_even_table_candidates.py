#!/usr/bin/env python3
"""Fill empty Paper 1-5 slots for even computational tables with curated papers."""

from __future__ import annotations

import argparse
import json
import re
import sys
import time
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


CANDIDATES: dict[int, list[str]] = {
    54: [
        "10.1038/emboj.2011.286",
        "10.1016/j.cell.2005.06.043",
        "10.1083/jcb.200911018",
        "10.1038/ncb2000",
        "10.1242/jcs.128868",
    ],
    55: [
        "10.1126/science.1153124",
        "10.1111/j.1600-0854.2009.00920.x",
        "10.1038/ncb2502",
        "10.1242/jcs.128868",
        "10.1016/j.devcel.2011.07.016",
    ],
    56: [
        "10.1038/ncb2000",
        "10.1242/jcs.115.12.2505",
        "10.1083/jcb.200911018",
        "10.1083/jcb.201710132",
        "10.1038/ncomms13588",
    ],
    57: [
        "10.1083/jcb.201710132",
        "10.1038/ncb2000",
        "10.1074/jbc.m301642200",
        "10.1083/jcb.200911018",
        "10.1038/ncomms13588",
    ],
    58: [
        "10.1016/j.cub.2009.09.059",
        "10.1016/j.tcb.2008.11.003",
        "10.1073/pnas.1200448109",
        "10.1073/pnas.1817498116",
        "10.1073/pnas.2417145122",
    ],
    59: [
        "10.1016/j.cub.2009.09.059",
        "10.1016/j.tcb.2008.11.003",
        "10.1073/pnas.1200448109",
        "10.1038/s41418-019-0342-5",
        "10.1083/jcb.202408159",
    ],
    60: [
        "10.1038/s41418-019-0342-5",
        "10.1038/nature09413",
        "10.1007/s00018-021-04078-0",
        "10.1002/jev2.12365",
        "10.1038/s41419-021-04317-z",
    ],
    67: [
        "10.1016/j.cell.2019.02.029",
        "10.1038/s41556-018-0040-4",
        "10.1038/s41556-021-00693-y",
        "10.1080/20013078.2020.1757209",
        "10.1073/pnas.1521230113",
    ],
    68: [
        "10.1038/ncomms3980",
        "10.7554/elife.19276",
        "10.7554/elife.71982",
        "10.1016/j.celrep.2016.09.031",
        "10.1093/nar/gkw1284",
    ],
    69: [
        "10.1038/nri3622",
        "10.1111/imm.13471",
        "10.1016/j.ymthe.2023.10.021",
        "10.1002/advs.202308662",
        "10.1038/85438",
    ],
    70: [
        "10.1101/cshperspect.a041415",
        "10.1016/j.bbalip.2013.04.011",
        "10.1194/jlr.R084210",
        "10.1126/science.1153124",
        "10.1002/jev2.12233",
    ],
    71: [
        "10.1080/20013078.2018.1535750",
        "10.1002/jev2.12404",
        "10.1038/nature15756",
        "10.1126/sciadv.adh1168",
        "10.15252/embj.201696003",
    ],
    72: [
        "10.1016/j.jconrel.2013.08.014",
        "10.1016/j.ab.2013.12.001",
        "10.1016/j.jconrel.2014.11.029",
        "10.1016/j.ymthe.2018.05.024",
        "10.1016/j.jconrel.2015.03.033",
    ],
    79: [
        "10.1371/journal.pbio.0060299",
        "10.1038/nature08012",
        "10.1038/s41418-019-0342-5",
        "10.1016/j.abb.2012.08.004",
        "10.1002/jev2.12365",
    ],
    80: [
        "10.1371/journal.pbio.0060299",
        "10.1038/nature08012",
        "10.1083/jcb.200404158",
        "10.1074/jbc.M404893200",
        "10.1016/j.cell.2019.02.029",
    ],
    81: [
        "10.1038/nchembio711",
        "10.1038/nrm2970",
        "10.1038/s41419-021-04317-z",
        "10.1002/jev2.12365",
        "10.1016/j.cell.2019.02.029",
    ],
    82: [
        "10.1080/20013078.2018.1535750",
        "10.1016/j.cell.2019.02.029",
        "10.1038/s41556-018-0040-4",
        "10.1038/s41556-021-00805-8",
        "10.1038/s41598-020-57497-7",
    ],
    83: [
        "10.3402/jev.v2i0.19861",
        "10.1080/20013078.2018.1535750",
        "10.1016/j.cell.2019.02.029",
        "10.1038/srep17319",
        "10.1038/s41598-020-57497-7",
    ],
    84: [
        "10.3390/ijms241914999",
        "10.1038/s41598-022-04868-x",
        "10.1002/jev2.12404",
        "10.1002/biot.201800528",
        "10.1080/20013078.2019.1609206",
    ],
    91: [
        "10.1080/20013078.2018.1535750",
        "10.1002/jev2.12404",
        "10.1002/biot.201800528",
        "10.1080/20013078.2018.1442088",
        "10.1016/j.jcyt.2025.06.003",
    ],
    92: [
        "10.1002/biot.201800528",
        "10.1172/jci.insight.99263",
        "10.1080/20013078.2018.1442088",
        "10.1172/jci.insight.99263",
        "10.1016/j.jlb.2024.100278",
    ],
    93: [
        "10.1038/s41598-020-57497-7",
        "10.3402/jev.v2i0.19861",
        "10.1038/srep17319",
        "10.1080/20013078.2018.1535750",
        "10.1002/jev2.12404",
    ],
    94: [
        "10.3402/jev.v2i0.19861",
        "10.1080/20013078.2018.1535750",
        "10.1002/jev2.12404",
        "10.1016/j.cell.2019.02.029",
        "10.1038/s41598-020-57497-7",
    ],
    95: [
        "10.1080/20013078.2018.1535750",
        "10.1080/20013078.2019.1609206",
        "10.1002/biot.201800528",
        "10.1016/j.jlb.2024.100278",
        "10.1002/jev2.12404",
    ],
    96: [
        "10.3390/ijms241914999",
        "10.1038/s41598-022-04868-x",
        "10.1002/biot.201800528",
        "10.1080/20013078.2019.1609206",
        "10.1002/jev2.12404",
    ],
}


FALLBACK: dict[str, str] = {
    "10.1073/pnas.1521230113": "Kowal, J., Arras, G., Colombo, M., Jouve, M., Morath, J. P., Primdal-Bengtson, B., Dingli, F., Loew, D., Tkach, M., & Thery, C. (2016). Proteomic comparison defines novel markers to characterize heterogeneous populations of extracellular vesicle subtypes. Proceedings of the National Academy of Sciences, 113(8), E968-E977. https://doi.org/10.1073/pnas.1521230113",
}


def clean_title(title: str) -> str:
    title = re.sub(r"<[^>]+>", "", title)
    return re.sub(r"\s+", " ", title).strip().rstrip(".")


def author_text(authors: list[dict[str, str]], max_authors: int = 6) -> str:
    names = []
    for author in authors[:max_authors]:
        family = author.get("family") or ""
        given = author.get("given") or ""
        initials = " ".join(part[0] + "." for part in re.split(r"[\s-]+", given) if part)
        names.append((family + (", " + initials if initials else "")).strip(", "))
    if not names:
        return ""
    if len(authors) > max_authors:
        names.append("et al.")
    if len(names) == 1:
        return names[0]
    return ", ".join(names[:-1]) + ", & " + names[-1]


def year(meta: dict) -> str:
    for key in ("published-print", "published-online", "published", "issued"):
        parts = meta.get(key, {}).get("date-parts", [])
        if parts and parts[0]:
            return str(parts[0][0])
    return "n.d."


def citation_from_doi(doi: str) -> str:
    doi = doi.lower()
    if doi in FALLBACK:
        return FALLBACK[doi]
    url = f"https://doi.org/{urllib.parse.quote(doi)}"
    req = urllib.request.Request(
        url,
        headers={
            "Accept": "application/vnd.citationstyles.csl+json",
            "User-Agent": "codex-even-table-curator/1.0",
        },
    )
    with urllib.request.urlopen(req, timeout=30) as response:
        meta = json.loads(response.read().decode("utf-8"))
    authors = author_text(meta.get("author", []))
    title = clean_title((meta.get("title") or [""])[0] if isinstance(meta.get("title"), list) else meta.get("title", ""))
    container = meta.get("container-title") or ""
    if isinstance(container, list):
        container = container[0] if container else ""
    volume = meta.get("volume") or ""
    issue = meta.get("issue") or ""
    page = meta.get("page") or meta.get("article-number") or ""
    vol_issue = volume + (f"({issue})" if issue else "")
    tail = ", ".join(x for x in [container, vol_issue, page] if x)
    return f"{authors} ({year(meta)}). {title}. {tail}. https://doi.org/{doi}"


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    wb = load_workbook(args.workbook)
    ws = wb["Literature Tracker"]
    report = []

    cache: dict[str, str] = {}
    for row, dois in CANDIDATES.items():
        table = ws.cell(row, 1).value
        if table not in {"A10", "A12", "A14", "A16"}:
            raise RuntimeError(f"row {row} is {table}, not an editable target table")
        for offset, doi in enumerate(dois, start=12):
            cell = ws.cell(row, offset)
            if cell.value not in (None, ""):
                continue
            if doi not in cache:
                try:
                    cache[doi] = citation_from_doi(doi)
                    time.sleep(0.15)
                except Exception as exc:
                    print(f"SKIP {doi}: {exc}", file=sys.stderr)
                    continue
            if not args.dry_run:
                cell.value = cache[doi]
            report.append({"cell": cell.coordinate, "table": table, "doi": doi, "citation": cache[doi]})

    if not args.dry_run:
        wb.save(args.workbook)

    print(json.dumps(report, indent=2, ensure_ascii=False))
    print(f"updated_cells={len(report)}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
