#!/usr/bin/env python3
"""Apply verified corrections to previously filled even-table paper cells."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook

from curate_even_table_candidates import citation_from_doi


CORRECTIONS = {
    "N72": "10.1016/j.jconrel.2014.11.029",
    "O72": "10.1016/j.ymthe.2018.05.024",
    "O92": "10.1016/j.jcyt.2025.06.003",
}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True, type=Path)
    args = parser.parse_args()

    wb = load_workbook(args.workbook)
    ws = wb["Literature Tracker"]
    report = []
    for cell, doi in CORRECTIONS.items():
        citation = citation_from_doi(doi)
        ws[cell].value = citation
        report.append({"cell": cell, "doi": doi, "citation": citation})
    wb.save(args.workbook)
    print(json.dumps(report, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
